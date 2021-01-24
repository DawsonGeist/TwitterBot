import ast

from TwitterAPI import TwitterAPI as t
import requests
from requests_oauthlib import OAuth1
import word_finder
import pprint
import xlsxwriter
import openpyxl
import time
import webbrowser


#   Twitter API Keys
twitterAPIKEY = ''
twitterSecret = ''
twitterBearer = ''
accessAPIKey = ''
accessSecret = ''

#   oauth bs
oauth_nonce = ''
oauth_signature_method = ''
oauth_timestamp = ''
oauth_version = '1.0'
oauth_signature = ''



#   Headers
headers = {'Authorization': 'Bearer ' + str(twitterBearer)}
headers2 = {
    'authorization': 'OAuth oauth_consumer_key=' + str(twitterAPIKEY) + ', oauth_nonce='+str(oauth_nonce)+', oauth_signature='+str(oauth_signature)+', oauth_signature_method="HMAC-SHA1", oauth_timestamp='+str(oauth_timestamp)+', oauth_token=' + str(accessAPIKey) + ', oauth_version="1.0"',
    'content-type': 'application/json'
}

#   Twitter Accounts
targets = [
    'Froste',
    'Nadeshot',
    'JhbTeam',
    'Class',
    'Viperous',
    'Avalanche100T'
]

#   Twitter Accounts 2
targets2 = [
    'jschlatt',
    'shroud',
    'REALMizkif',
    'ConnorEatsPants',
    'xoxabstract',
    'KaleiRenay'
]

#   Generate API Object
api = t(twitterAPIKEY, twitterSecret, accessAPIKey, accessSecret)



#   Tweet Something
def tweet(msg):
    r = api.request('statuses/update', {'status': str(msg)})
    print(r.text)


#   Reply to Tweet
def tweetReply(id, msg):
    r = api.request('statuses/update', {'status': str(msg), 'in_reply_to_status_id': str(id)})
    return r


# https://developer.twitter.com/en/docs/twitter-api/v1/tweets/search/api-reference/get-search-tweets
def searchForTweets(keyword):
    r = api.request('search/tweets', {'q': str(keyword)})
    for item in r:
        pprint.pprint(item)
        print('\n\n\n Next Entity \n\n\n')


def getUserId(username):
    r = requests.get('https://api.twitter.com/2/users/by/username/' + str(username), headers=headers)
    print(r.json())
    return r.json()

#   Rate of 20 requests per minute (300 every 15 mins)
def getUserByID(id):
    r = requests.get('https://api.twitter.com/2/users/' + str(id), headers=headers)
    time.sleep(3)
    try:
        print(r.json())
        return r.json()
    except:
        print('JSON ERROR ' + str(r.text))


def getReplysToUser(userID):
    r = requests.get('https://api.twitter.com/2/users/'+str(userID)+'/tweets?expansions=in_reply_to_user_id&max_results=50',headers=headers)
    time.sleep(3)
    pprint.pprint(r.json())
    data = r.json()
    users = set()
    usersIDs = set()
    returner = []
    for tweet in data['data']:
        if 'in_reply_to_user_id' in tweet:
            id = tweet['in_reply_to_user_id']
            user = getUserByID(id)
            if user != None:
                try:
                    #       THIS SHIT IS BROKEN, ID != USERNAME --- SETS ARE UNORDERED
                    name = user['data']['username']
                    uID = user['data']['id']
                    print("Interacted with " + name)
                    users.add(name)
                    usersIDs.add(uID)
                except:
                    print('user suspended')
            else:
                pass
    returner.append(users)
    returner.append(usersIDs)
    return returner


#   https://developer.twitter.com/en/docs/twitter-api/tweets/timelines/api-reference/get-users-id-tweets
def getTweetsByUser(userID, excludes, maxResults, tweetFields, mediaFields, userFields, pagination_token):
    if(maxResults < 5):
        maxResults = 5

    if(pagination_token != ''):
        params = (('tweet.fields', tweetFields),
                  ('media.fields', mediaFields),
                  ('user.fields', userFields),
                  ('exclude', excludes),
                  ('pagination_token', pagination_token),
                  ('max_results', str(maxResults)))
    else:
        params = (('tweet.fields', tweetFields),
                  ('media.fields', mediaFields),
                  ('user.fields', userFields),
                  ('exclude', excludes),
                  ('max_results', str(maxResults)))

    auth = OAuth1(twitterAPIKEY, twitterSecret, accessAPIKey, accessSecret)
    r = requests.get(
        'https://api.twitter.com/2/users/' + str(userID) + '/tweets?',
        params=params, headers=headers, auth=auth)
    print(r.status_code)
    pprint.pprint(r.json())
    return r.json()


#   App does not have write action - Email Twitter
def followUserByID(id):
    url = "https://api.twitter.com/1.1/friendships/create.json?user_id="+str(id)+"&follow=true"
    #   God bless this fucking function
    auth = OAuth1(twitterAPIKEY, twitterSecret, accessAPIKey, accessSecret)
    response = requests.request("POST", url, auth=auth)
    time.sleep(3)
    #   Disable Notifications
    url2 = "https://api.twitter.com/1.1/friendships/update.json?user_id="+str(id)+"&device=false"
    auth2 = OAuth1(twitterAPIKEY, twitterSecret, accessAPIKey, accessSecret)
    response2 = requests.request("POST", url2, auth=auth2)
    time.sleep(3)
    pprint.pprint(response.json())
    print('\n\n')
    pprint.pprint(response2.json())



def getGlobalTweetStream():
    # Not allowed to share
    return False

#   1 request a minute
def getFollowers(userID, maxResults):
    r = requests.get('https://api.twitter.com/2/users/'+str(userID)+'/followers?max_results='+str(maxResults),
                     headers=headers)
    time.sleep(60)
    try:
        print(r.json())
        return r.json()
    except:
        print('JSON ERROR ' + str(r.text))

#   Access Denied Email Twitter
def sendTextDM(userID, msg):
    auth = OAuth1(twitterAPIKEY, twitterSecret, accessAPIKey, accessSecret)

    headers = {
        'content-type': 'application/json'
    }

    data = '{"event": {"type": "message_create", "message_create": {"target": {"recipient_id": "' + str(userID)\
           + '"}, "message_data": {"text": "' + str(msg)+'"}}}}'

    response = requests.post('https://api.twitter.com/1.1/direct_messages/events/new.json', headers=headers,
                             data=data, auth=auth)
    print(response.json())


def junkWork0():
    workbook = xlsxwriter.Workbook('Target_Accounts_v2.xlsx')
    worksheet = workbook.add_worksheet()
    count = 0
    for users in targets2:
        temp = getUserId(str(users))
        print(temp['data']['id'])
        worksheet.write(count, 0, targets2[count])
        worksheet.write(count, 1, '\"' + temp['data']['id'] + '\"')
        count += 1
    workbook.close()


def junkWork1():
    index = 0
    workbook = xlsxwriter.Workbook('Targets_Responded_To2.xlsx')
    targetWorkbook = openpyxl.load_workbook('Target_Accounts_v2.xlsx')
    sheet = targetWorkbook['Sheet1']

    for x in range(len(targets2)):
        userID = str(sheet.cell(x+1, 2).value).replace('\"', '')
        spokeToo = getReplysToUser(userID)
        worksheet = workbook.add_worksheet(sheet.cell(x+1, 1).value)
        for i in range(len(spokeToo[0])):
            worksheet.write(index, 0, spokeToo[0].pop())
            worksheet.write(index, 1, ('\"' + spokeToo[1].pop()) + '\"')
            index += 1
        index = 0
    workbook.close()

def junkWork2():
    index = 1
    workbook = openpyxl.load_workbook('Targets_Responded_To2.xlsx')
    sheets = workbook.sheetnames
    for sheet in sheets:
        page = workbook[sheet]
        while(page.cell(index, 2).value != None):
            id = page.cell(index, 2).value
            id = str(id).replace('\"', '')
            followUserByID(id)
            index += 1
        index = 1


def junkWork3():
    user = '1040361509396905987'
    # retweets,replies
    exclude = 'retweets,replies'
    maxResults = 5
    mediaFields = 'public_metrics'
    tweetFields = 'id,text,public_metrics'
    userFields = 'public_metrics,verified'

    t = getTweetsByUser(userID=user, excludes=exclude, maxResults=maxResults, mediaFields=mediaFields,
                        tweetFields=tweetFields, userFields=userFields, pagination_token='')
    token = t['meta']['next_token']

    t = getTweetsByUser(userID=user, excludes=exclude, maxResults=maxResults, mediaFields=mediaFields,
                        tweetFields=tweetFields, userFields=userFields, pagination_token=token)


def twitter_user_sign_in():
    url = 'https://api.twitter.com/oauth/request_token'
    auth = OAuth1(twitterAPIKEY, twitterSecret, accessAPIKey, accessSecret)
    header = {
        'oauth_callback': 'oob'
    }
    r = requests.request('POST', url=url, auth=auth, headers=header)
    temp = []
    if(r.status_code != 200):
        print('Bad Request: ' + str(r.text))
        return ''
    else:
        body = r.text
        for i in range(3):
            index = body.rfind('=')
            value = body[index+1:]
            temp.append(value)
            body = body[:body.rfind('&')]
        print(temp)
    print(r.text)
    user_oauth_token = temp[2]
    user_oauth_secret = temp[1]

    webbrowser.open('https://api.twitter.com/oauth/authorize?oauth_token=' + str(temp[2]))
    # print(r2.status_code)
    # print(r2.text)



# tweetReply('1343692543658053633', 'Test')
# junkWork0()
# junkWork1()
# junkWork2()
twitter_user_sign_in()

# Open AWS account for redirect URL so users can authenticate Twitter use
# within app
